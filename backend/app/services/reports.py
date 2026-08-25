"""Generate downloadable PDF and Excel reports from live database data."""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.supply_chain import Invoice, InvoiceStatus
from app.models.vendoriq import ComplianceDocument, Contract, PurchaseOrder
from app.services.analytics import compute_spend_over_time
from app.services.performance import compute_all_vendors_performance
from app.services.reliability import compute_vendor_ranking

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
PDF_HEADER_BG = colors.HexColor("#1E3A5F")
PDF_ALT_ROW = colors.HexColor("#F1F5F9")


def _now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def _fmt_money(value) -> str:
    return f"${float(value or 0):,.2f}"


def _fmt_pct(value) -> str:
    if value is None:
        return "—"
    return f"{float(value):.1f}%"


def _fmt_date(value) -> str:
    if value is None:
        return "—"
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).strftime("%Y-%m-%d")


def _fmt_hours(value) -> str:
    if value is None:
        return "—"
    return f"{float(value):.1f}h"


# ---------------------------------------------------------------------------
# Data collectors
# ---------------------------------------------------------------------------


def collect_vendor_performance(db: Session) -> dict:
    ranking = compute_vendor_ranking(db)
    performance = {item.vendor_id: item for item in compute_all_vendors_performance(db)}
    headers = [
        "Rank",
        "Vendor",
        "Reliability",
        "Risk",
        "On-Time %",
        "Quality",
        "Completion %",
        "Avg Response",
    ]
    rows = []
    for entry in ranking:
        metrics = performance.get(entry.vendor_id)
        rows.append(
            [
                entry.rank,
                entry.vendor_name,
                f"{entry.overall_score:.1f}",
                entry.risk_level,
                _fmt_pct(metrics.on_time_delivery_pct if metrics else None),
                _fmt_pct(metrics.average_quality_score if metrics else None),
                _fmt_pct(metrics.order_completion_rate if metrics else None),
                _fmt_hours(metrics.average_response_time_hours if metrics else None),
            ]
        )

    risk_counts = Counter(entry.risk_level for entry in ranking)
    kpis = [
        ("Vendors scored", str(len(ranking))),
        ("Low risk", str(risk_counts.get("Low", 0))),
        ("Medium risk", str(risk_counts.get("Medium", 0))),
        ("High risk", str(risk_counts.get("High", 0))),
    ]
    return {
        "title": "Vendor Performance Report",
        "filename": f"vendor-performance-{_now_stamp()}",
        "kpis": kpis,
        "sheets": [{"name": "Vendor Performance", "headers": headers, "rows": rows}],
    }


def collect_procurement_summary(db: Session) -> dict:
    orders = list(db.scalars(select(PurchaseOrder).options(selectinload(PurchaseOrder.vendor))))
    spend = compute_spend_over_time(db)
    status_counts = Counter(str(order.status) for order in orders)
    total_value = sum(float(order.total_amount or 0) for order in orders)

    summary_headers = ["Metric", "Value"]
    summary_rows = [
        ["Total purchase orders", len(orders)],
        ["Total spend (non-cancelled)", _fmt_money(spend.total_spend)],
        ["All-status PO value", _fmt_money(total_value)],
        ["Monthly periods", len(spend.points)],
    ]
    for status, count in sorted(status_counts.items()):
        summary_rows.append([f"POs — {status}", count])

    monthly_headers = ["Period", "Spend", "Order count"]
    monthly_rows = [
        [point.period, _fmt_money(point.total_spend), point.order_count]
        for point in spend.points
    ]

    po_headers = ["PO Number", "Vendor", "Status", "Order date", "Amount", "Currency"]
    po_rows = [
        [
            order.po_number,
            order.vendor.name if order.vendor else "—",
            str(order.status),
            _fmt_date(order.order_date),
            _fmt_money(order.total_amount),
            order.currency,
        ]
        for order in sorted(orders, key=lambda item: item.order_date, reverse=True)[:500]
    ]

    kpis = [
        ("Purchase orders", str(len(orders))),
        ("Total spend", _fmt_money(spend.total_spend)),
        ("Statuses tracked", str(len(status_counts))),
    ]
    return {
        "title": "Procurement Summary Report",
        "filename": f"procurement-summary-{_now_stamp()}",
        "kpis": kpis,
        "sheets": [
            {"name": "Summary", "headers": summary_headers, "rows": summary_rows},
            {"name": "Spend Over Time", "headers": monthly_headers, "rows": monthly_rows},
            {"name": "Purchase Orders", "headers": po_headers, "rows": po_rows},
        ],
    }


def collect_compliance(db: Session) -> dict:
    contracts = list(
        db.scalars(select(Contract).options(selectinload(Contract.vendor)).order_by(Contract.expiry_date))
    )
    documents = list(
        db.scalars(
            select(ComplianceDocument).options(selectinload(ComplianceDocument.vendor))
        )
    )
    invoices = list(db.scalars(select(Invoice)))

    flag_counts = Counter(str(contract.compliance_flag) for contract in contracts)
    status_counts = Counter(str(contract.status) for contract in contracts)
    overdue = sum(1 for invoice in invoices if str(invoice.status) == InvoiceStatus.OVERDUE.value)

    summary_headers = ["Metric", "Value"]
    summary_rows = [
        ["Contracts", len(contracts)],
        ["Compliance documents", len(documents)],
        ["Overdue invoices", overdue],
    ]
    for flag, count in sorted(flag_counts.items()):
        summary_rows.append([f"Contracts — {flag}", count])
    for status, count in sorted(status_counts.items()):
        summary_rows.append([f"Status — {status}", count])

    contract_headers = [
        "Contract #",
        "Vendor",
        "Title",
        "Status",
        "Compliance",
        "Start",
        "Expiry",
        "Value",
    ]
    contract_rows = [
        [
            contract.contract_number,
            contract.vendor.name if contract.vendor else "—",
            contract.title,
            str(contract.status),
            str(contract.compliance_flag),
            _fmt_date(contract.start_date),
            _fmt_date(contract.expiry_date),
            _fmt_money(contract.contract_value),
        ]
        for contract in contracts
    ]

    doc_headers = ["Vendor", "Type", "Name", "Status", "Uploaded", "Expires"]
    doc_rows = [
        [
            document.vendor.name if document.vendor else "—",
            document.document_type,
            document.document_name,
            document.status,
            _fmt_date(document.uploaded_at),
            _fmt_date(document.expires_at),
        ]
        for document in documents
    ]

    kpis = [
        ("Contracts", str(len(contracts))),
        ("Compliant", str(flag_counts.get("Compliant", 0))),
        ("Non-compliant", str(flag_counts.get("Non-Compliant", 0))),
        ("Overdue invoices", str(overdue)),
    ]
    return {
        "title": "Compliance Report",
        "filename": f"compliance-{_now_stamp()}",
        "kpis": kpis,
        "sheets": [
            {"name": "Summary", "headers": summary_headers, "rows": summary_rows},
            {"name": "Contracts", "headers": contract_headers, "rows": contract_rows},
            {"name": "Documents", "headers": doc_headers, "rows": doc_rows},
        ],
    }


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------


def render_excel(payload: dict) -> bytes:
    workbook = Workbook()
    first = True
    for sheet in payload["sheets"]:
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = sheet["name"][:31]
        worksheet["A1"] = payload["title"]
        worksheet["A1"].font = TITLE_FONT
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(sheet["headers"]), 2))
        worksheet["A2"] = f"Generated {_now_stamp()} (UTC) from live VendorIQ data"
        worksheet["A2"].font = Font(italic=True, color="64748B")

        start_row = 4
        if payload.get("kpis") and sheet is payload["sheets"][0]:
            worksheet["A4"] = "Highlights"
            worksheet["A4"].font = Font(bold=True)
            for index, (label, value) in enumerate(payload["kpis"], start=5):
                worksheet.cell(index, 1, label)
                worksheet.cell(index, 2, value)
            start_row = 5 + len(payload["kpis"]) + 1

        header_row = start_row
        for col, header in enumerate(sheet["headers"], start=1):
            cell = worksheet.cell(header_row, col, header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        rows = sheet["rows"] or [["No records found"] + [""] * (len(sheet["headers"]) - 1)]
        for row_index, row in enumerate(rows, start=header_row + 1):
            for col, value in enumerate(row, start=1):
                worksheet.cell(row_index, col, value)

        for col in range(1, len(sheet["headers"]) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_pdf(payload: dict) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        title=payload["title"],
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        textColor=PDF_HEADER_BG,
        fontSize=16,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["Normal"],
        textColor=colors.HexColor("#64748B"),
        fontSize=9,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        textColor=PDF_HEADER_BG,
        fontSize=12,
        spaceBefore=10,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)

    story = [
        Paragraph(payload["title"], title_style),
        Paragraph(f"Generated {_now_stamp()} (UTC) from live VendorIQ data", meta_style),
    ]

    if payload.get("kpis"):
        kpi_data = [[Paragraph("<b>Highlight</b>", cell_style), Paragraph("<b>Value</b>", cell_style)]]
        kpi_data.extend(
            [Paragraph(str(label), cell_style), Paragraph(str(value), cell_style)]
            for label, value in payload["kpis"]
        )
        kpi_table = Table(kpi_data, colWidths=[3.2 * inch, 2.4 * inch])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), PDF_HEADER_BG),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#EFF6FF")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 12))

    usable_width = landscape(letter)[0] - inch
    for sheet in payload["sheets"]:
        story.append(Paragraph(sheet["name"], section_style))
        headers = [Paragraph(f"<b>{header}</b>", cell_style) for header in sheet["headers"]]
        body_rows = sheet["rows"] or [["No records found"] + [""] * (len(sheet["headers"]) - 1)]
        data = [headers]
        for row in body_rows[:80]:
            data.append([Paragraph(str(value), cell_style) for value in row])

        col_width = usable_width / max(len(sheet["headers"]), 1)
        table = Table(data, colWidths=[col_width] * len(sheet["headers"]), repeatRows=1)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), PDF_HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for index in range(1, len(data)):
            if index % 2 == 0:
                style_commands.append(("BACKGROUND", (0, index), (-1, index), PDF_ALT_ROW))
        table.setStyle(TableStyle(style_commands))
        story.append(table)
        if len(body_rows) > 80:
            story.append(Paragraph(f"Showing first 80 of {len(body_rows)} rows. Use Excel for the full export.", meta_style))

    document.build(story)
    return buffer.getvalue()


def build_report(payload: dict, fmt: str) -> tuple[bytes, str, str]:
    if fmt == "xlsx":
        content = render_excel(payload)
        return content, f"{payload['filename']}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    content = render_pdf(payload)
    return content, f"{payload['filename']}.pdf", "application/pdf"
